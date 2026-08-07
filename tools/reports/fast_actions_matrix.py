# tools/reports/fast_actions_matrix.py
import os
from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook

# Config
DATABASE_URL = os.getenv("SUPABASE_DATABASE_URL") or os.getenv("DATABASE_URL")
TABLE = os.getenv("FAST_TABLE", "fast_actions")           # ajuste se necessário
COL_AGENT = os.getenv("FAST_AGENT_COL", "agent_name")
COL_ACTION = os.getenv("FAST_ACTION_COL", "action")
COL_DATE = os.getenv("FAST_DATE_COL", "created_at")
OUTDIR = os.getenv("REPORT_OUT", "./reports")
UPLOAD_TO_DRIVE = os.getenv("UPLOAD_REPORT_TO_DRIVE", "false").lower() in ("1","true","yes")
DRIVE_REPORT_FOLDER_ID = os.getenv("DRIVE_REPORT_FOLDER_ID")  # opcional, se upload ativado

os.makedirs(OUTDIR, exist_ok=True)

def normalize(s):
    if pd.isna(s):
        return s
    return ' '.join(str(s).upper().strip().split())

def fetch_data():
    if not DATABASE_URL:
        raise RuntimeError("Defina SUPABASE_DATABASE_URL no ambiente")
    engine = create_engine(DATABASE_URL)
    sql = f"""
        SELECT
          {COL_AGENT} AS agent_name,
          {COL_ACTION} AS action,
          {COL_DATE}::timestamp AT TIME ZONE 'UTC' AS created_at
        FROM {TABLE}
        WHERE {COL_ACTION} IS NOT NULL
    """
    df = pd.read_sql(sql, engine)
    if df.empty:
        return df
    df['agent_name'] = df['agent_name'].apply(normalize)
    df['action'] = df['action'].astype(str).str.strip()
    df['year'] = df['created_at'].dt.year
    df['month'] = df['created_at'].dt.month
    df['ym'] = df['created_at'].dt.strftime('%Y-%m')
    return df

def build_reports(df):
    timestamp = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    out_xlsx = os.path.join(OUTDIR, f"fast_actions_matrix_{timestamp}.xlsx")

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Sheet: Data (long table) with filters
    ws_data = wb.create_sheet("Data")
    df_sort = df.sort_values(['agent_name','ym','action'])
    for r in dataframe_to_rows(df_sort, index=False, header=True):
        ws_data.append(r)
    # set autofilter
    ws_data.auto_filter.ref = f"A1:{chr(65 + len(df_sort.columns)-1)}{len(df_sort)+1}"

    # Unique actions (for columns)
    actions = sorted(df['action'].dropna().unique())

    # Build a "Matrix_ALL" (agents x actions) aggregated (whether agent ever did action)
    agents = sorted(df['agent_name'].dropna().unique())
    matrix_all = pd.DataFrame(0, index=agents, columns=actions)
    grouped = df.groupby(['agent_name','action']).size().reset_index(name='cnt')
    for _, row in grouped.iterrows():
        matrix_all.at[row['agent_name'], row['action']] = 1

    # write Matrix_ALL
    ws_all = wb.create_sheet("Matrix_ALL")
    ws_all.append(["Agent"] + actions)
    for agent in matrix_all.index:
        ws_all.append([agent] + matrix_all.loc[agent].astype(int).tolist())

    # Create monthly sheets: Matrix_YYYY-MM
    months = sorted(df['ym'].unique())
    for ym in months:
        dfm = df[df['ym']==ym]
        mat = pd.DataFrame(0, index=sorted(dfm['agent_name'].unique()), columns=actions)
        g = dfm.groupby(['agent_name','action']).size().reset_index(name='cnt')
        for _, row in g.iterrows():
            mat.at[row['agent_name'], row['action']] = 1
        ws = wb.create_sheet(f"Matrix_{ym}")
        ws.append(["Agent"] + actions)
        for agent in mat.index:
            ws.append([agent] + mat.loc[agent].astype(int).tolist())

    # Apply conditional formatting colors (green for 1, red for 0) to all matrix sheets
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # light red

    for sheet_name in [s.title for s in wb.worksheets if s.title.startswith('Matrix_') or s.title=='Matrix_ALL']:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        # iterate cells except header column 1
        for r in range(2, max_row+1):
            for c in range(2, max_col+1):
                cell = ws.cell(row=r, column=c)
                try:
                    v = int(cell.value or 0)
                except:
                    v = 0
                if v >= 1:
                    cell.fill = green
                    cell.value = ""  # keep cell visually just colored; you can change to 'X' if prefer
                else:
                    cell.fill = red
                    cell.value = ""

    # Create Index sheet with links to monthly sheets
    ws_idx = wb.create_sheet("Índice")
    ws_idx.append(["Available months"])
    for ym in months:
        ws_idx.append([ym])
        # create hyperlink to sheet
        link_cell = ws_idx.cell(row=ws_idx.max_row, column=1)
        link_cell.hyperlink = f"#{'Matrix_'+ym}!A1"
        link_cell.style = "Hyperlink"

    # Save
    wb.save(out_xlsx)
    return out_xlsx

def upload_to_drive(filepath, folder_id):
    # Optional: upload using googleapiclient; requires service account json in env variable file and drive API enabled.
    # Keep implementation here minimal; call only if UPLOAD_TO_DRIVE True and folder_id provided.
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except Exception as e:
        print("google client libs not available:", e)
        return None

    sa_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE")
    if not sa_file:
        print("NO service account file configured for Drive upload.")
        return None

    creds = service_account.Credentials.from_service_account_file(sa_file, scopes=['https://www.googleapis.com/auth/drive.file'])
    drive = build('drive','v3', credentials=creds, cache_discovery=False)
    file_metadata = {'name': os.path.basename(filepath)}
    if folder_id:
        file_metadata['parents'] = [folder_id]
    media = MediaFileUpload(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    f = drive.files().create(body=file_metadata, media_body=media, fields='id,webViewLink').execute()
    return f

def main():
    df = fetch_data()
    if df.empty:
        print("Nenhum dado retornado da consulta; relatório não gerado.")
        return
    out = build_reports(df)
    print("Relatório salvo em:", out)
    if UPLOAD_TO_DRIVE and DRIVE_REPORT_FOLDER_ID:
        res = upload_to_drive(out, DRIVE_REPORT_FOLDER_ID)
        print("Upload result:", res)

if __name__ == '__main__':
    main()
